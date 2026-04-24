"""
조달청 나라장터 발주계획현황 - 용역 월별 수집기
================================================

공공데이터포털의 "조달청_나라장터 발주계획현황서비스"를 호출해
지정한 연도·월의 용역 발주계획을 수집하고,
관리시트 템플릿(A~S 열, 19열)과 동일한 형식의 엑셀로 저장한다.

사용법
------
  python3 research.py <연도2자리> <월 ...>

예시
----
  python3 research.py 26 3         → 2026년 3월만
  python3 research.py 26 1 2 3     → 2026년 1·2·3월
  python3 research.py 25 6         → 2025년 6월만
"""

from __future__ import annotations

import os
import sys
import time
from datetime import datetime
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# =========================================================================
# 1. 환경 설정
# =========================================================================

SERVICE_KEY: str = os.environ.get("DATA_GO_KR_KEY", "")

# API 설정
BASE_URL = "https://apis.data.go.kr/1230000/ao/OrderPlanSttusService"
OPERATION = "getOrderPlanSttusListServc"
PAGE_SIZE = 999
REQUEST_SLEEP_SEC = 0.05
TIMEOUT_SEC = 20
MAX_RETRY = 3


# =========================================================================
# 2. API 호출부
# =========================================================================

def fetch_page(page_no: int, order_bgn_ym: str, order_end_ym: str,
               inqry_bgn_dt: str, inqry_end_dt: str) -> dict[str, Any]:
    """한 페이지 조회. JSON으로 반환받음."""
    params = {
        "ServiceKey": SERVICE_KEY,
        "type": "json",
        "numOfRows": PAGE_SIZE,
        "pageNo": page_no,
        "inqryDiv": 1,
        "orderBgnYm": order_bgn_ym,
        "orderEndYm": order_end_ym,
        "inqryBgnDt": inqry_bgn_dt,
        "inqryEndDt": inqry_end_dt,
    }
    url = f"{BASE_URL}/{OPERATION}"

    last_err: Exception | None = None
    for attempt in range(1, MAX_RETRY + 1):
        try:
            resp = requests.get(url, params=params, timeout=TIMEOUT_SEC)
            resp.raise_for_status()
            try:
                data = resp.json()
            except ValueError:
                raise RuntimeError(f"비JSON 응답: {resp.text[:500]}")
            return data
        except Exception as e:
            last_err = e
            wait = 2 ** attempt
            print(f"[page {page_no}] 오류: {e} -> {wait}초 후 재시도 ({attempt}/{MAX_RETRY})")
            time.sleep(wait)
    raise RuntimeError(f"page {page_no} 조회 실패: {last_err}")


def extract_items_and_total(data: dict[str, Any]) -> tuple[list[dict], int]:
    """응답 구조에서 items 리스트와 totalCount 추출."""
    # nkoneps 에러 응답 구조 처리 (예: 입력범위값 초과 에러)
    nk_error = data.get("nkoneps.com.response.ResponseError")
    if nk_error:
        err_header = nk_error.get("header", {})
        err_code = err_header.get("resultCode", "")
        err_msg = err_header.get("resultMsg", "")
        raise RuntimeError(f"API 오류 [{err_code}] {err_msg}")

    body = data.get("response", {}).get("body", {})
    header = data.get("response", {}).get("header", {})

    result_code = str(header.get("resultCode", ""))
    result_msg = header.get("resultMsg", "")
    if result_code and result_code != "00":
        if result_code != "03":
            raise RuntimeError(f"API 오류 [{result_code}] {result_msg}")
        return [], 0

    total = int(body.get("totalCount") or 0)
    raw_items = body.get("items")

    if isinstance(raw_items, dict):
        inner = raw_items.get("item", [])
        items = inner if isinstance(inner, list) else ([inner] if inner else [])
    elif isinstance(raw_items, list):
        items = raw_items
    else:
        items = []
    return items, total


def fetch_all_for_month(year: int, month: int) -> list[dict]:
    """특정 연도·월의 전체 용역 발주계획 수집."""
    ym = f"{year:04d}{month:02d}"
    
    # API의 게시일 범위 제한(최대 1년)을 우회하면서 항상 최신 데이터를 얻기 위해
    # 조회 시작일을 '현재 시점 기준 1년 전'으로, 종료일을 '현재 시점'으로 설정합니다.
    now = datetime.now()
    inqry_end = now.strftime("%Y%m%d%H%M")
    
    # 약 365일 전
    from datetime import timedelta
    inqry_bgn = (now - timedelta(days=365)).strftime("%Y%m%d%H%M")

    print(f"[{month}월] 조회 범위: 발주년월 {ym}, 게시일 {inqry_bgn}~{inqry_end}")
    first = fetch_page(1, ym, ym, inqry_bgn, inqry_end)
    items, total = extract_items_and_total(first)
    print(f"[{month}월] -> 전체 {total}건 예상, 페이지당 {PAGE_SIZE}건")

    collected = list(items)
    total_pages = (total + PAGE_SIZE - 1) // PAGE_SIZE if total else 1

    for p in range(2, total_pages + 1):
        time.sleep(REQUEST_SLEEP_SEC)
        data = fetch_page(p, ym, ym, inqry_bgn, inqry_end)
        page_items, _ = extract_items_and_total(data)
        collected.extend(page_items)
        print(f"[{month}월] -> page {p}/{total_pages} 누적 {len(collected)}")

    print(f"[{month}월] 수집 완료: {len(collected)}건")
    return collected


# =========================================================================
# 3. 관리시트 템플릿 매핑 (A~S, 19열)
# =========================================================================

# 헤더 (행2) - 관리시트 템플릿 그대로
HEADERS = [
    "투찰 준비",                    # A
    "No.",                          # B
    "용역명",                       # C
    "수요기관명",                    # D
    "기관구분",                      # E
    "지역",                         # F
    "계약방법",                      # G
    "발주시기(원본)",                # H
    "정확한 공고시기 확인\n(유선 조사)",  # I
    "발주금액(원)",                  # J
    "우선순위",                      # K
    "전담팀원 (에이블런)",            # L
    "진행상태",                      # M
    "전략메모",                      # N
    "발주계획번호",                  # O
    "조달구분",                      # P
    "담당부서",                      # Q
    "담당자",                        # R
    "연락처",                        # S
]

# 컬럼 폭 (관리시트 템플릿 기준)
COL_WIDTHS = {
    "A": 5.83, "B": 4.66, "C": 60.16, "D": 13.0, "E": 13.0,
    "F": 24.66, "G": 13.0, "H": 13.0, "I": 13.0, "J": 13.0,
    "K": 13.0, "L": 13.0, "M": 13.0, "N": 13.0, "O": 13.0,
    "P": 13.0, "Q": 13.0, "R": 13.0, "S": 13.0,
}


def map_row(it: dict, row_num: int) -> list:
    """API item -> 관리시트 1행 (A~S) 으로 변환."""
    # 발주시기를 YYYY/MM 문자열 형식으로 변환
    order_year = str(it.get("orderYear") or "").strip()
    order_mnth = str(it.get("orderMnth") or "").strip().zfill(2)
    order_ym = f"{order_year}/{order_mnth}" if order_year and order_mnth.isdigit() else ""

    # 발주금액
    amt_raw = it.get("sumOrderAmt")
    try:
        amt = int(str(amt_raw).replace(",", "")) if amt_raw not in (None, "") else None
    except ValueError:
        amt = None

    return [
        None,                                              # A: 투찰 준비 (수동)
        row_num,                                           # B: No.
        (it.get("bizNm") or "").strip(),                   # C: 용역명
        (it.get("orderInsttNm") or "").strip(),            # D: 수요기관명
        (it.get("jrsdctnDivNm") or "").strip(),            # E: 기관구분
        (it.get("cnstwkRgnNm") or "").strip(),             # F: 지역
        (it.get("cntrctMthdNm") or "").strip(),            # G: 계약방법
        order_ym,                                          # H: 발주시기(원본)
        None,                                              # I: 정확한 공고시기 확인 (수동)
        amt,                                               # J: 발주금액(원)
        None,                                              # K: 우선순위 (수동)
        None,                                              # L: 전담팀원 (수동)
        None,                                              # M: 진행상태 (수동)
        None,                                              # N: 전략메모 (수동)
        (it.get("orderPlanUntyNo") or "").strip(),          # O: 발주계획번호
        (it.get("prcrmntMethd") or "").strip(),             # P: 조달구분
        (it.get("deptNm") or "").strip(),                   # Q: 담당부서
        (it.get("ofclNm") or "").strip(),                   # R: 담당자
        (it.get("telNo") or "").strip(),                    # S: 연락처
    ]


# =========================================================================
# 4. 엑셀 저장 (관리시트 템플릿 동일 서식)
# =========================================================================

def write_xlsx(month_data: dict[int, list[dict]], year: int, path: str) -> None:
    """월별 데이터를 시트로 나누어 관리시트 형식으로 저장."""
    wb = Workbook()
    # 기본 시트 삭제
    wb.remove(wb.active)

    # 스타일 정의
    title_font = Font(name="Arial", size=12, bold=True)
    title_fill = PatternFill(start_color="FFB7B7B7", end_color="FFB7B7B7", fill_type="solid")
    title_align = Alignment(horizontal="left", vertical="center")

    header_font = Font(name="Noto Sans", size=9, bold=True)
    header_fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Malgun Gothic", size=9)
    data_font_arial = Font(name="Arial", size=9)

    for month in sorted(month_data.keys()):
        items = month_data[month]
        tab_name = f"{month}월 발주계획 마스터"
        ws = wb.create_sheet(title=tab_name)

        # ── 행1: 제목 (A1:S1 병합) ──
        ws.merge_cells("A1:S1")
        title_text = f"📋 {month}월 발주계획 마스터"
        ws["A1"] = title_text
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = title_align
        ws.row_dimensions[1].height = 23.25

        # ── 행2: 헤더 ──
        for i, h in enumerate(HEADERS, start=1):
            c = ws.cell(row=2, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
        ws.row_dimensions[2].height = 39.0

        # ── 행3 이후: 데이터 (발주금액 내림차순 정렬) ──
        # 먼저 금액 기준으로 정렬
        items.sort(key=lambda it: int(str(it.get("sumOrderAmt") or "0").replace(",", "") or "0"), reverse=True)

        for idx, it in enumerate(items, start=1):
            row = map_row(it, idx)
            row_num = idx + 2  # 행3부터 데이터 시작

            for col_idx, val in enumerate(row, start=1):
                c = ws.cell(row=row_num, column=col_idx, value=val)

                # 컬럼별 정렬 & 폰트 매핑 (템플릿 기준)
                if col_idx in (1, 2, 8, 11, 12, 13):
                    # A, B, H, K, L, M: Arial, center
                    c.font = data_font_arial
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 7:
                    # G: 계약방법 center
                    c.font = data_font
                    c.alignment = Alignment(horizontal="center")
                elif col_idx == 10:
                    # J: 발주금액 right, 숫자 서식
                    c.font = data_font
                    c.alignment = Alignment(horizontal="right")
                    if isinstance(val, (int, float)):
                        c.number_format = "#,##0"
                elif col_idx in (9, 14):
                    # I, N: left + vcenter
                    c.font = data_font_arial
                    c.alignment = Alignment(horizontal="left", vertical="center")
                elif col_idx == 16:
                    # P: 조달구분 center
                    c.font = data_font
                    c.alignment = Alignment(horizontal="center")
                elif col_idx == 18:
                    # R: 담당자 center
                    c.font = data_font
                    c.alignment = Alignment(horizontal="center")
                else:
                    # C, D, E, F, O, Q, S: left
                    c.font = data_font
                    c.alignment = Alignment(horizontal="left")

        # ── 컬럼 폭 ──
        for col, w in COL_WIDTHS.items():
            ws.column_dimensions[col].width = w

        # ── 틀 고정 (행2 헤더 아래) ──
        ws.freeze_panes = "A3"

        print(f"  [{tab_name}] {len(items)}건 기록")

    wb.save(path)
    print(f"\n저장 완료: {path}")


# =========================================================================
# 5. 진입점
# =========================================================================

def print_usage():
    print("사용법: python3 research.py <연도2자리> <월 ...>")
    print("  예) python3 research.py 26 3       → 2026년 3월")
    print("  예) python3 research.py 26 1 2 3   → 2026년 1·2·3월")
    print("  예) python3 research.py 25 6       → 2025년 6월")


def main() -> int:
    if len(sys.argv) < 3:
        print_usage()
        return 1

    if SERVICE_KEY == "여기에_디코딩된_인증키_입력" or not SERVICE_KEY:
        print("ERROR: SERVICE_KEY 가 설정되지 않았습니다.")
        return 1

    # 인자 파싱
    try:
        year_short = int(sys.argv[1])
        year = 2000 + year_short
        months = sorted(set(int(m) for m in sys.argv[2:]))
        for m in months:
            if not 1 <= m <= 12:
                raise ValueError(f"잘못된 월: {m}")
    except ValueError as e:
        print(f"입력 오류: {e}")
        print_usage()
        return 1

    month_label = ", ".join(f"{m}월" for m in months)
    print(f"=== {year}년 {month_label} 용역 발주계획 수집 시작 ===\n")

    # 월별 수집
    month_data: dict[int, list[dict]] = {}
    for m in months:
        try:
            items = fetch_all_for_month(year, m)
            month_data[m] = items
        except Exception as e:
            print(f"[{m}월] 수집 중단: {e}")
            return 2
        print()

    if not any(month_data.values()):
        print("수집된 데이터가 없습니다. 날짜 범위나 인증키 승인 상태를 확인하세요.")

    # 파일명 생성
    os.makedirs("outputs", exist_ok=True)
    if len(months) == 1:
        fname = os.path.join("outputs", f"{year}년_{months[0]}월_용역_발주계획.xlsx")
    else:
        fname = os.path.join("outputs", f"{year}년_{months[0]}-{months[-1]}월_용역_발주계획.xlsx")

    write_xlsx(month_data, year, fname)
    return 0


if __name__ == "__main__":
    sys.exit(main())